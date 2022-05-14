from schedules.models import Schedule,Report
from aggregates.models import DataLockdate
from schedules.views import search_staff_tr_query
from staffs.models import User
from hana.mixins import StaffUserRequiredMixin,SuperUserRequiredMixin,jpweek
from django.views.generic import TemplateView,ListView,View
from django.http import HttpResponse,Http404,HttpResponseRedirect
from django.utils.timezone import make_aware,localtime
from django.db.models import Q
from django.urls import reverse
import json
import datetime
import calendar
import math
import openpyxl
import re

from dateutil.relativedelta import relativedelta


class TopView(SuperUserRequiredMixin,TemplateView):
    model = Schedule
    template_name = "aggregates/aggregate_top.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        now = make_aware(datetime.datetime.now())
        context['time_now'] = now
        context['month_before'] = now - relativedelta(months=1)

        data_lock_date = DataLockdate.objects.first()
        context['data_lock_date'] = localtime(data_lock_date.lock_date) if data_lock_date else None

        #画面推移後の戻るボタン用にpathをセッションに記録
        self.request.session['from'] = self.request.get_full_path()

        return context

class DataLockView(SuperUserRequiredMixin,View):    
    def get(self,request, **kwargs):
        if not self.request.GET.get('year') or not self.request.GET.get('month'):
            raise Http404
        year = int(self.request.GET.get('year'))
        month= int(self.request.GET.get('month'))
        lock_date = make_aware(datetime.datetime(year,month,1) + relativedelta(months=1) - datetime.timedelta(seconds=1))

        #テーブルを空にする
        DataLockdate.objects.all().delete()
        #追加
        DataLockdate.objects.get_or_create(lock_date=lock_date,updated_by=self.request.user)
        url = reverse('aggregates:aggregate_top')
        return HttpResponseRedirect(url)

class KaigoView(SuperUserRequiredMixin,ListView):
    model = Schedule
    template_name = "aggregates/invoice_kaigo.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        year = self.kwargs.get('year')
        month= self.kwargs.get('month')
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)
        next_month     = this_month + relativedelta(months=1)
        before_month   = this_month - relativedelta(months=1)

        context['this_month']    = this_month
        context['next_month']    = next_month
        context['before_month']  = before_month

        queryset = Schedule.objects.select_related('report','careuser','service').filter(service__kind=0,report__careuser_confirmed=True,\
                   report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('careuser__last_kana','careuser__first_kana','report__service_in_date')

        cu = kaigo_list(queryset)        
        context['careuser']  = cu

        return context

def kaigo_export(request,year,month):
    
    if request.user.is_superuser:
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)

        queryset = Schedule.objects.select_related('report','careuser','service').filter(service__kind=0,report__careuser_confirmed=True,\
                    report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('careuser__last_kana','careuser__first_kana','report__service_in_date')
        cu_data = kaigo_list(queryset)
        
        json_data =  json.dumps(cu_data,ensure_ascii=False)

        response = HttpResponse(json_data,content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="kaigo.json"'

        return response
    else:
        return Http404

def kaigo_list(schedules):
    cu={}
    for sche in schedules:
        #careuser毎のリスト作成
        careuser_name = sche.careuser.last_name + " " + sche.careuser.first_name
        if careuser_name not in cu:
            cu[careuser_name]=[]

        s_in_time  = localtime(sche.report.service_in_date)
        s_out_time = localtime(sche.report.service_out_date)

        #スケジュールをサービス内容と日時毎に分類
        service  = sche.service.bill_title
        mix_items = sche.service.mix_items
        night    = False
        midnight = False

        if   s_in_time.time() >= datetime.time(18,00) and s_in_time.time() < datetime.time(22,00):
            night = True
        elif s_in_time.time() >= datetime.time(6,00) and s_in_time.time() < datetime.time(8,00):
            night = True
        elif s_in_time.time() >= datetime.time(0,00) and s_in_time.time() < datetime.time(6,00):
            midnight = True
        elif s_in_time.time() >= datetime.time(22,00) and s_in_time.time() < datetime.time(23,59,59):
            midnight = True
        
        peoples = sche.peoples
        in_time  = str(s_in_time.hour).zfill(2) + ":" + str(s_in_time.minute).zfill(2)
        out_time = str(s_out_time.hour).zfill(2) + ":" + str(s_out_time.minute).zfill(2)        

        add_check = False

        #同日のスケジュールがあれば日付を追記
        if cu[careuser_name]:
            for sche in cu[careuser_name]:
                if sche['service'] == service and sche['night']==night and sche['midnight']==midnight and sche['peoples']==peoples and sche['in_time'] == in_time and sche['out_time'] == out_time:
                   sche['date'].append(int(s_in_time.day))
                   add_check = True
                   break

        #なければ新たなスケジュールを作成
        if not add_check:
            new_obj = {}
            new_obj['year']      = s_in_time.strftime("%Y")
            new_obj['month']     = s_in_time.strftime("%m")
            new_obj['service']   = service
            new_obj['mix_items'] = mix_items
            new_obj['night']     = night
            new_obj['midnight']  = midnight
            new_obj['peoples']   = peoples
            new_obj['in_time']   = in_time
            new_obj['out_time']  = out_time
            new_obj['date']      = [int(s_in_time.day)]

            cu[careuser_name].append(new_obj)

    return cu

class SougouView(SuperUserRequiredMixin,ListView):
    model = Schedule
    template_name = "aggregates/invoice_sougou.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        year = self.kwargs.get('year')
        month= self.kwargs.get('month')
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)
        next_month     = this_month + relativedelta(months=1)
        before_month   = this_month - relativedelta(months=1)

        context['this_month']    = this_month
        context['next_month']    = next_month
        context['before_month']  = before_month

        queryset = Schedule.objects.select_related('report','careuser','service').filter(service__kind=3,report__careuser_confirmed=True,\
                   report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('careuser__last_kana','careuser__first_kana','report__service_in_date')

        cu = kaigo_list(queryset)        
        context['careuser']  = cu

        return context

def sougou_export(request,year,month):
    
    if request.user.is_superuser:
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)

        queryset = Schedule.objects.select_related('report','careuser','service').filter(service__kind=3,report__careuser_confirmed=True,\
                    report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('careuser__last_kana','careuser__first_kana','report__service_in_date')
        
        cu_data = kaigo_list(queryset)        
        json_data =  json.dumps(cu_data,ensure_ascii=False)

        response = HttpResponse(json_data,content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="sougou.json"'

        return response
    else:
        return Http404

class InvoiceView(SuperUserRequiredMixin,ListView):
    model = Schedule
    template_name = "aggregates/invoice.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        kind = self.kwargs.get('kind')
        context['kind'] = kind
        
        #障害のみ編集
        if kind == "1-0":
            kind = 1
            genre_query = (Q(service__bill_title__contains="身体") | Q(service__bill_title__contains="家事") | Q(service__bill_title__contains="通院"))
        elif kind == "1-1":
            kind = 1
            genre_query = Q(service__bill_title__contains="重度")
        else:
            kind = int(kind)
            genre_query = Q()


        year = self.kwargs.get('year')
        month= self.kwargs.get('month')
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)
        next_month     = this_month + relativedelta(months=1)
        before_month   = this_month - relativedelta(months=1)

        context['this_month']    = this_month
        context['next_month']    = next_month
        context['before_month']  = before_month

        queryset = Schedule.objects.select_related('report','careuser','service').filter(genre_query,service__kind=kind,report__careuser_confirmed=True,\
                   report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('careuser__last_kana','careuser__first_kana','report__service_in_date')

        context['data']  = export_list(queryset,kind)

        return context

def export(request,kind,year,month):
    
    if request.user.is_superuser:
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)

        #障害のみ編集
        if kind == "1-0":
            serch_kind = 1
            genre_query = (Q(service__bill_title__contains="身体") | Q(service__bill_title__contains="家事") | Q(service__bill_title__contains="通院"))
        elif kind == "1-1":
            serch_kind = 1
            genre_query = Q(service__bill_title__contains="重度")
        else:
            serch_kind = int(kind)
            genre_query = Q()

        queryset = Schedule.objects.select_related('report','careuser','service','staff1','staff2','staff3','staff4').filter(genre_query,service__kind=serch_kind,report__careuser_confirmed=True,\
                    report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('careuser__last_kana','careuser__first_kana','report__service_in_date')

        cu_data = export_list(queryset,kind) 
        
        json_data =  json.dumps(cu_data,ensure_ascii=False)

        response = HttpResponse(json_data,content_type='application/json')
        if kind == "1-0":
            response['Content-Disposition'] = 'attachment; filename="shougai_kyotaku.json"'
        if kind == "1-1":
            response['Content-Disposition'] = 'attachment; filename="shougai_juudo.json"'
        if kind == "4":
            response['Content-Disposition'] = 'attachment; filename="shougai_doukou.json"'
        elif kind == "2":
            response['Content-Disposition'] = 'attachment; filename="idou.json"'
        elif kind == "5":
            response['Content-Disposition'] = 'attachment; filename="jihi.json"'

        return response
    else:
        return Http404

def json_serial(obj):
    # 日付型の場合には、文字列に変換します
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()
    # 上記以外はサポート対象外.
    #raise TypeError ("Type %s not serializable" % type(obj))

def export_list(schedules,kind):
    cu={}
    for sche in schedules:

        #careuser毎のリスト作成
        careuser_name = sche.careuser.last_name + " " + sche.careuser.first_name
        if careuser_name not in cu:
            cu[careuser_name]=[]

        #複合の場合はバラシてリスト化
        #混合でない場合
        if not sche.service.mix_items:            
            sv = sche.service.bill_title
            in_datetime  = localtime(sche.report.service_in_date)
            out_datetime = localtime(sche.report.service_out_date)

            new_obj = set_dict(sche,sv,in_datetime,out_datetime)
            #開始時間または終了時間がlist内のスケジュールと一致したら、スケジュールの時間を変更する
            #一致しなければ追加する
            if not same_time_joint(cu[careuser_name],new_obj):cu[careuser_name].append(new_obj)

        #混合の場合
        else:
            if sche.report.mix_reverse == False:
                sv1 = sche.service.name_main + str(sche.service.in_time_main)
                in_datetime1 = localtime(sche.report.service_in_date)
                out_datetime1 = in_datetime1 + datetime.timedelta(minutes=sche.report.in_time_main)

                sv2 = sche.service.name_sub + str(sche.service.in_time_sub)
                in_datetime2 = out_datetime1
                out_datetime2 = localtime(sche.report.service_out_date)

            else:
                sv1 = sche.service.name_sub + str(sche.service.in_time_sub)
                in_datetime1 = localtime(sche.report.service_in_date)
                out_datetime1 = in_datetime1 + datetime.timedelta(minutes=sche.report.in_time_sub)

                sv2 = sche.service.name_main + str(sche.service.in_time_main)
                in_datetime2 = out_datetime1
                out_datetime2 = localtime(sche.report.service_out_date)

            new_obj = set_dict(sche,sv1,in_datetime1,out_datetime1)
            #開始時間または終了時間がlist内のスケジュールと一致したら、スケジュールの時間を変更する
            #一致しなければ追加する
            if not same_time_joint(cu[careuser_name],new_obj):cu[careuser_name].append(new_obj)
                
            #二つ目を追加
            new_obj = set_dict(sche,sv2,in_datetime2,out_datetime2)
            #開始時間または終了時間がlist内のスケジュールと一致したら、スケジュールの時間を変更する
            #一致しなければ追加する
            if not same_time_joint(cu[careuser_name],new_obj):cu[careuser_name].append(new_obj)
      
    return cu

def set_dict(sche,srv,in_datetime,out_datetime):
    obj = {}

    obj['in_year']       = in_datetime.year
    obj['in_month']      = in_datetime.month
    obj['in_day']        = in_datetime.day
    obj['in_hour']       = in_datetime.hour
    obj['in_minute']     = in_datetime.minute
    obj['in_time']       = str(in_datetime.hour).zfill(2) + ":" + str(in_datetime.minute).zfill(2)
    obj['night']         = night_check(in_datetime)
    obj['midnight']      = midnight_check(in_datetime)

    obj['out_year']      = out_datetime.year
    obj['out_month']     = out_datetime.month
    obj['out_day']       = out_datetime.day
    obj['out_hour']      = out_datetime.hour
    obj['out_minute']    = out_datetime.minute
    obj['out_time']      = str(out_datetime.hour).zfill(2) + ":" + str(out_datetime.minute).zfill(2)

    obj['service']       = srv
    if obj['night']   :obj['service'] += "<夜間>"
    if obj['midnight']:obj['service'] += "<深夜>"

    obj['mix_items'] = sche.service.mix_items                
    obj['peoples']   = sche.peoples
    obj['staff1']    = sche.staff1.get_short_name() if sche.staff1 else None
    obj['staff2']    = sche.staff2.get_short_name() if sche.staff2 else None
    obj['staff3']    = sche.staff3.get_short_name() if sche.staff3 else None
    obj['staff4']    = sche.staff4.get_short_name() if sche.staff4 else None
    obj['biko']      = sche.biko if sche.biko else ""
    if sche.report.communicate:
        if obj['biko']:obj['biko'] += "　"
        obj['biko'] += sche.report.communicate
    obj['error']     = sche.report.get_error_code_display() if sche.report.error_code else ""
    obj['warnings']  = sche.report.warnings if sche.report.warnings else ""
    obj['adding']    = False#合算
    
    return obj

def night_check(in_datetime):
    ret = False
    if in_datetime.time() >= datetime.time(18,00) and in_datetime.time() < datetime.time(22,00):
        ret = True
    elif in_datetime.time() >= datetime.time(6,00) and in_datetime.time() < datetime.time(8,00):
        ret = True
    return ret

def midnight_check(in_datetime):
    ret = False
    if in_datetime.time() >= datetime.time(0,00) and in_datetime.time() < datetime.time(6,00):
        ret = True
    elif in_datetime.time() >= datetime.time(22,00) and in_datetime.time() < datetime.time(23,59,59):
        ret = True
    return ret

def same_time_joint(obj,add_dict):
    add_check = False
    for s in obj:
        srv       = ''.join([i for i in s['service'] if not i.isdigit()]).replace('<夜間>','').replace('<深夜>','')#数字を除去
        add_srv = ''.join([i for i in add_dict['service'] if not i.isdigit()]).replace('<夜間>','').replace('<深夜>','')#数字を除去
        in_datetime      = make_aware(datetime.datetime(s['in_year'],s['in_month'],s['in_day'],s['in_hour'],s['in_minute']))
        add_in_datetime  = make_aware(datetime.datetime(add_dict['in_year'],add_dict['in_month'],add_dict['in_day'],add_dict['in_hour'],add_dict['in_minute']))
        out_datetime     = make_aware(datetime.datetime(s['out_year'],s['out_month'],s['out_day'],s['out_hour'],s['out_minute']))
        add_out_datetime = make_aware(datetime.datetime(add_dict['out_year'],add_dict['out_month'],add_dict['out_day'],add_dict['out_hour'],add_dict['out_minute']))

        if in_datetime.date() == add_in_datetime.date() and srv == add_srv and s['peoples'] == add_dict['peoples'] and (in_datetime == add_out_datetime or out_datetime == add_in_datetime):
            s['service'] += " + " + add_dict['service']
            if in_datetime == add_out_datetime:
                s['in_year']    = add_dict['in_year']
                s['in_month']   = add_dict['in_month']
                s['in_day']     = add_dict['in_day']
                s['in_hour']    = add_dict['in_hour']
                s['in_minute']  = add_dict['in_minute']
                s['in_time']    = add_dict['in_time']
                s['night']      = night_check(add_in_datetime)
                s['midnight']   = midnight_check(add_in_datetime)
            elif out_datetime == add_in_datetime:
                s['out_year']    = add_dict['out_year']
                s['out_month']   = add_dict['out_month']
                s['out_day']     = add_dict['out_day']
                s['out_hour']    = add_dict['out_hour']
                s['out_minute']  = add_dict['out_minute']
                s['out_time']    = add_dict['out_time']
                s['night']       = night_check(in_datetime)
                s['midnight']    = midnight_check(in_datetime)
            
            s['adding']   = True#合算

            if s['staff1'] != add_dict['staff1']:s['staff1'] = s['staff1'] + "・" + add_dict['staff1']
            if s['staff2'] != add_dict['staff2']:s['staff2'] = s['staff2'] + "・" + add_dict['staff2']
            if s['staff3'] != add_dict['staff3']:s['staff3'] = s['staff3'] + "・" + add_dict['staff3']
            if s['staff4'] != add_dict['staff4']:s['staff4'] = s['staff4'] + "・" + add_dict['staff4']
            
            if add_dict['biko']:
                if s['biko']:s['biko'] += "　"
                s['biko'] += add_dict['biko']
            if add_dict['error']:
                if s['error']:s['error'] += "　"
                s['error'] += add_dict['error']
            if add_dict['warnings']:
                if s['warnings']:s['warnings'] += "　"
                s['warnings'] += add_dict['warnings']
            
            add_check = True
            break
    return add_check


class SalaryEmployeeView(SuperUserRequiredMixin,ListView):
    model = Schedule
    template_name = "aggregates/salaryemployee.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        year = self.kwargs.get('year')
        month= self.kwargs.get('month')
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)
        next_month     = this_month + relativedelta(months=1)
        before_month   = this_month - relativedelta(months=1)

        context['this_month']    = this_month
        context['next_month']    = next_month
        context['before_month']  = before_month

        #スタッフ毎の実績を取得
        staffs = User.objects.filter(salary=1).order_by('-is_staff','last_kana','first_kana')
        staff_obj_list = []
        for staff in staffs:
            obj = {}
            obj['staff'] = staff

            queryset = Schedule.objects.select_related('report','careuser','service').filter(search_staff_tr_query(staff),report__careuser_confirmed=True,\
                       report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('report__service_in_date')
            
            obj['schedules'] = queryset
            staff_obj_list.append(obj)

        #給与出力用にlistを生成
        achieve = salalyemployee_achieve_list(staff_obj_list,year,month)        
        context['achieve']  = achieve

        return context

def salalyemployee_export(request,year,month):
    
    if request.user.is_superuser:
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)

        staffs = User.objects.filter(salary=1).order_by('-is_staff','last_kana','first_kana')
        staff_obj_list = []
        for staff in staffs:
            obj = {}
            obj['staff'] = staff

            condition_staff = search_staff_tr_query(staff)
            queryset = Schedule.objects.select_related('report','careuser','service').filter(condition_staff,report__careuser_confirmed=True,\
                       report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('report__service_in_date')
            
            if queryset:
                obj['schedules'] = queryset
                staff_obj_list.append(obj)

        #給与出力用ファイル生成
        achieve = salalyemployee_achieve_list(staff_obj_list,year,month)

        #wb = openpyxl.load_workbook('aggregates/monthly_employee.xlsx')
        wb = openpyxl.Workbook()
        sheet = wb.active        

        #罫線
        side   = openpyxl.styles.borders.Side(style='thin', color='000000')
        border = openpyxl.styles.borders.Border(top=side, bottom=side, left=side, right=side)
        #背景色
        fill   = openpyxl.styles.PatternFill(patternType='solid', fgColor='d3d3d3')
        fill_for_input = openpyxl.styles.PatternFill(patternType='solid', fgColor='FFFF00')

        sheet_name = "R" + str(year-2018) + "." + str(month)

        #シートの存在を確認
        is_sheet = False
        for ws in wb.worksheets:
            if ws.title == sheet_name:
                is_sheet = True
                break
        
        font = openpyxl.styles.Font(name='BIZ UDゴシック')

        #シートが存在していなければ作成
        if not is_sheet: 
            wb.create_sheet(title=sheet_name,index=0)
            ws = wb[sheet_name]
            ws.sheet_view.showGridLines = False #目盛り線を消す
            
            ws['A1'] = '岸田さんについては泊りの場合は時間外加算は計上不要。泊りでない場合は時間外を計上（2022/01より）'
            ws.merge_cells('A1:K1')
            ws['A2'] = '22:00~以降は時間外加算を支給'
            ws.merge_cells('A1:K1')

            print_start = "A3"
            row_height = 24

            #列の幅を調整
            ws.column_dimensions['A'].width = 2
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 22
            ws.column_dimensions['G'].width = 6.5
            ws.column_dimensions['H'].width = 8
            ws.column_dimensions['I'].width = 8
            ws.column_dimensions['J'].width = 14
            ws.column_dimensions['k'].width = 12.5

            for staff_data in achieve:
                row = ws.max_row + 3
                ws.row_dimensions[row].height = 30 #行の高さ
                ws.cell(row,2,value=str(year) + "年" + str(month) + "月  " + staff_data['staff_name'] + " 様")
                ws.cell(row,2).font = openpyxl.styles.fonts.Font(size=16)
                ws.cell(row,2).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
                row +=1
                ws.row_dimensions[row].height = row_height #行の高さ
                ws.cell(row,2,value="日付")
                ws.cell(row,3,value="時間(分)")
                ws.cell(row,4,value="実施分数")
                ws.cell(row,5,value="利用者")
                ws.cell(row,6,value="サービス")
                ws.cell(row,7,value="研修")
                ws.cell(row,8,value="適用時間")
                ws.cell(row,9,value="移動時間")
                ws.cell(row,10,value="22-5時間外加算")
                ws.cell(row,11,value="合計時間")

                #センターリング・罫線・背景色
                for r in  ws.iter_rows(min_row=row, min_col=2, max_row=row, max_col=11):
                    for c in r:
                        c.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                        ws[c.coordinate].border = border
                        ws[c.coordinate].fill   = fill

                index = row+1
                start_row = index #合計値計算用
                for day,data in staff_data['days_data'].items():
                    if data['schedules']:
                        day_start_row = index
                        day_end_row   = index + len(data['schedules'])-1
                        ws.cell(index,2,value= str(day) + "日(" + data['week'] + ")")
                        ws.cell(index,2).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                        #結合
                        ws.merge_cells(ws.cell(row=index,column=2).coordinate + ":" + ws.cell(row=index+len(data['schedules'])-1,column=2).coordinate)
                        ws.cell(index,11,value='=SUM(' + ws.cell(row=day_start_row,column=8).coordinate + ':' + ws.cell(row=day_end_row,column=9).coordinate + ')')
                        ws.cell(index,11).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                        #結合
                        ws.merge_cells(ws.cell(row=index,column=11).coordinate + ":" + ws.cell(row=index+len(data['schedules'])-1,column=11).coordinate)
                        
                        for sche in data['schedules']:
                            ws.row_dimensions[index].height = row_height #行の高さ
                            ws.cell(index,3,value=sche['s_in_time'] + "～" + sche['s_out_time'])
                            ws.cell(index,3).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                            ws.cell(index,4,value=sche['real_minutes'])
                            ws.cell(index,4).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                            ws.cell(index,5,value=sche['careuser']+" 様")
                            ws.cell(index,5).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                            ws.cell(index,6,value=sche['service'])
                            ws.cell(index,6).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                            if sche['kenshuu']:
                                ws.cell(index,7,value="[研修]")
                            ws.cell(index,7).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                            ws.cell(index,8,value=sche['adopt_hour'])
                            ws.cell(index,8).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                            if sche['move_hour'] > 0:ws.cell(index,9,value=sche['move_hour'])
                            ws.cell(index,9).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                            if sche['over_night'] : ws.cell(index,10,value='泊り')
                            elif sche['off_hour'] > 0: ws.cell(index,10,value=sche['off_hour'])
                            ws.cell(index,10).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                            index += 1

                end_row = index-1 #合計値計算用
                #罫線
                for r in  ws.iter_rows(min_row=start_row, min_col=2, max_row=end_row, max_col=11):
                    for c in r:
                        ws[c.coordinate].border = border

                row = ws.max_row+1 
                ws.cell(row,9,value='合計')
                ws.cell(row,9).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                ws.row_dimensions[row].height = row_height #行の高さ
                ws.cell(row,9).fill   = fill
                ws.cell(row,9).border = border
                ws.cell(row,9).border = border
                ws.cell(row,10,value='=SUM(' + ws.cell(row=start_row,column=10).coordinate + ':' + ws.cell(row=end_row,column=10).coordinate + ')')
                ws.cell(row,10).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                ws.cell(row,10).border = border
                ws.cell(row,11,value='=SUM(' + ws.cell(row=start_row,column=11).coordinate + ':' + ws.cell(row=end_row,column=11).coordinate + ')')
                ws.cell(row,11).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                ws.cell(row,11).border = border
                
                ws.cell(row+2,10,value='事務時間（分）')
                ws.cell(row+2,10).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                ws.row_dimensions[row+2].height = row_height #行の高さ
                ws.cell(row+2,10).fill   = fill
                ws.cell(row+2,10).border = border

                ws.cell(row+2,11).border = border
                ws.cell(row+2,11).fill   = fill_for_input
                ws.cell(row+2,11).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')

                ws.cell(row+3,10,value='泊り(回)')
                ws.row_dimensions[row+3].height = row_height #行の高さ
                ws.cell(row+3,10).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                ws.cell(row+3,10).fill   = fill
                ws.cell(row+3,10).border = border

                ws.cell(row+3,11).border = border
                ws.cell(row+3,11).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                ws.cell(row+3,11,value=staff_data['month_over_night'])

                ws.cell(row+4,10,value='総合計時間')
                ws.cell(row+4,10).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                ws.row_dimensions[row+4].height = row_height #行の高さ
                ws.cell(row+4,10).fill   = fill
                ws.cell(row+4,10).border = border

                ws.cell(row+4,11).border = border
                ws.cell(row+4,11,value='=' + ws.cell(row=row,column=11).coordinate + ' + FLOOR(' + ws.cell(row=row+2,column=11).coordinate + '/60,0.25)')
                ws.cell(row+4,11).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')


                #改ページ
                row = ws.max_row+2 
                page_break = openpyxl.worksheet.pagebreak.Break(id=row) # create Break obj 
                ws.page_breaks[0].append(page_break)

            #印刷範囲
            print_end = ws.cell(row=row,column=11).coordinate
            ws.print_area = print_start + ":" + print_end
            ws.page_setup.fitToWidth  = True
            ws.page_setup.fitToHeight = False
            ws.sheet_properties.pageSetUpPr.fitToPage = True

        #font
        #for row in ws:
        #    for cell in row:
        #        ws[cell.coordinate].font = font

        #出力
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=monthly_employee.xlsx'
        # データの書き込みを行なったExcelファイルを保存する
        #wb.save('aggregates/monthly_employee.xlsx')
        wb.save(response)

        # 生成したHttpResponseをreturnする
        return response 
    else:
        return Http404

def salalyemployee_achieve_list(staff_obj_list,year,month):

    archive = []

    for s in staff_obj_list:
        #一カ月の日数を取得(最終日のみ)
        days = calendar.monthrange(year,month)[1]
        obj ={}
        obj['staff_name'] = s['staff'].last_name + " " + s['staff'].first_name
        obj['month_total_hour'] = 0;
        obj['month_off_hour'] = 0;
        obj['month_over_night'] = 0;
        days_data = {}
        for day in range(days):
            # {1日のdatetime: 1日のスケジュール全て, 2日のdatetime: 2日の全て...}のような辞書を作る
            days_data[day+1] = {}
            days_data[day+1]['week'] = jpweek(make_aware(datetime.datetime(year,month,day+1,0,0)))
            days_data[day+1]['schedules'] = []
            days_data[day+1]['day_service_hour'] = 0
            days_data[day+1]['day_move_hour'] = 0     
            days_data[day+1]['off_hours'] = 0

        for sche in s['schedules']:
            d ={}
            s_in_date  = localtime(sche.report.service_in_date)
            s_out_date = localtime(sche.report.service_out_date)
            before_day =  (s_in_date - relativedelta(days=1)).day#前日の日付
            
            d['real_minutes']  = math.floor((s_out_date - s_in_date).total_seconds()/60)
            d['real_hour']     = math.ceil(d['real_minutes']/15)*0.25
            d['s_in_time']  = str(s_in_date.hour).zfill(2)  + ":" + str(s_in_date.minute).zfill(2)
            d['s_out_time'] = str(s_out_date.hour).zfill(2) + ":" + str(s_out_date.minute).zfill(2)
            d['s_in_time_datetime']  = s_in_date
            d['s_out_time_datetime'] = s_out_date

            
            d['careuser'] = sche.careuser.last_name + " " + sche.careuser.first_name
            d['service'] = ""
            if sche.service.kind==0:d['service'] += "[介護]"
            elif sche.service.kind==1:d['service'] += "[障害]"
            d['service']  += sche.service.title

            d['service_minutes']  = sche.service.time
            d['service_hour']     = math.ceil(sche.service.time/15)*0.25

            #同行チェック
            d['kenshuu'] = False
            if sche.tr_staff1 == s['staff'] or sche.tr_staff2 == s['staff'] or sche.tr_staff3 == s['staff'] or sche.tr_staff4 == s['staff']:
                d['kenshuu'] = True
            

            #実質時間または規定時間を計算適用時間とする。
            if d['real_hour'] > d['service_hour']:
                d['adopt_hour'] = d['real_hour']
            else:
                d['adopt_hour'] = d['real_hour']
            #合計時間に加算する。
            days_data[s_in_date.day]['day_service_hour'] += d['adopt_hour']
            obj['month_total_hour'] += d['adopt_hour']

            #備考に入力があれば付記。
            d['biko'] = ""
            if sche.biko:d['biko'] += sche.biko
            if sche.report.communicate:
                if d['biko']:d['biko'] += "　"
                d['biko'] += sche.report.communicate

            #移動時間を加算
            chk_flg = False
            d['move_hour'] = 0
            for sc in days_data[s_in_date.day]['schedules']:
                if sc['careuser'] == d['careuser']:
                     #全く同じ時間の場合
                    if s_in_date==sc['s_in_time_datetime'] and s_out_date == sc['s_out_time_datetime']:
                        chk_flg = True
                    #一部が重なる場合
                    elif s_in_date < sc['s_in_time_datetime'] and s_out_date > sc['s_in_time_datetime'] and s_out_date <= sc['s_out_time_datetime']:
                        chk_flg = True
                    elif s_in_date >= sc['s_in_time_datetime'] and s_in_date < sc['s_out_time_datetime'] and s_out_date > sc['s_out_time_datetime']:
                        chk_flg = True
                    #内包する場合
                    elif s_in_date < sc['s_in_time_datetime'] and s_out_date > sc['s_out_time_datetime']:
                        chk_flg = True
                    elif s_in_date > sc['s_in_time_datetime'] and s_out_date < sc['s_out_time_datetime']:
                        chk_flg = True
                    #繋がる予定の場合
                    elif s_in_date == sc['s_out_time_datetime'] or s_out_date == sc['s_in_time_datetime']:
                        chk_flg = True
            #前日と続くスケジュールをチェック
            if s_in_date.day > 1:
                if days_data[before_day]:
                    for sc in days_data[before_day]['schedules']:
                        if sc['careuser'] == d['careuser']:
                            #繋がる予定の場合
                            if s_in_date == sc['s_out_time_datetime']:
                                chk_flg = True
            if not chk_flg:
                d['move_hour'] = 0.25
                days_data[s_in_date.day]['day_move_hour'] += 0.25
                #合計時間に加算する。
                obj['month_total_hour'] += 0.25

            
            #22~5時の時間外時間を加算 #15分未満を切り捨てて時間変換（0.25を掛ける）にする
            oc130 = make_aware(datetime.datetime(s_in_date.year,s_in_date.month,s_in_date.day,1,30))
            oc5   = make_aware(datetime.datetime(s_in_date.year,s_in_date.month,s_in_date.day,5,0))
            oc22  = make_aware(datetime.datetime(s_in_date.year,s_in_date.month,s_in_date.day,22,0))
            d['off_hour'] = 0
            d['over_night'] = False#泊り

            #岸田さんの1:30以降は時間外を付けず、泊りとする。
            if sche.careuser.last_name == "岸田" and sche.careuser.first_name == "慶子" and s_in_date >= oc130 and s_in_date < oc5:
                check_over_night = False
                #泊りの重複を避ける
                for sc in days_data[s_in_date.day]['schedules']:
                    if sc['careuser'] == d['careuser'] and d['over_night']:
                        check_over_night = True
                if not check_over_night:
                    d['over_night'] = True
                    obj['month_over_night'] += 1
            #時間外の加算
            else:
                if s_in_date < oc5:
                    if s_out_date <= oc5: d['off_hour'] += math.floor(((s_out_date - s_in_date).total_seconds()/60)/15)*0.25
                    else: d['off_hour'] += math.floor(((oc5 - s_in_date).total_seconds()/60)/15)*0.25
                if s_out_date > oc22:
                    if s_in_date >= oc22: d['off_hour'] += math.floor(((s_out_date - s_in_date).total_seconds()/60)/15)*0.25
                    else: d['off_hour'] += math.floor(((s_out_date - oc22).total_seconds()/60)/15)*0.25

                days_data[s_in_date.day]['off_hours'] += d['off_hour']
                obj['month_off_hour'] += d['off_hour']

            #日毎のサービスと移動の合計時間を上書きセット
            days_data[s_in_date.day]['day_total_hour'] = days_data[s_in_date.day]['day_service_hour'] + days_data[s_in_date.day]['day_move_hour']

            days_data[s_in_date.day]['schedules'].append(d)

        obj['days_data'] = days_data
        archive.append(obj)

    return archive


class CommissionEmployeeView(SuperUserRequiredMixin,ListView):
    model = Schedule
    template_name = "aggregates/commissionemployee.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        year = self.kwargs.get('year')
        month= self.kwargs.get('month')
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)
        next_month     = this_month + relativedelta(months=1)
        before_month   = this_month - relativedelta(months=1)

        context['this_month']    = this_month
        context['next_month']    = next_month
        context['before_month']  = before_month

        #スタッフ毎の実績を取得
        staffs = User.objects.filter(salary=2).order_by('-is_staff','last_kana','first_kana')
        staff_obj_list = []
        for staff in staffs:
            obj = {}
            obj['staff'] = staff

            queryset = Schedule.objects.select_related('report','careuser','service').filter(search_staff_tr_query(staff),report__careuser_confirmed=True,\
                       report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('report__service_in_date')
            
            if queryset:
                obj['schedules'] = queryset
                staff_obj_list.append(obj)

        #給与出力用にlistを生成
        achieve = commissionemployee_achieve_list(staff_obj_list,year,month)        
        context['achieve']  = achieve

        return context

def commissionemployee_export(request,year,month):
    
    if request.user.is_superuser:
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)

        staffs = User.objects.filter(salary=2).order_by('-is_staff','last_kana','first_kana')
        staff_obj_list = []
        for staff in staffs:
            obj = {}
            obj['staff'] = staff

            queryset = Schedule.objects.select_related('report','careuser','service').filter(search_staff_tr_query(staff),report__careuser_confirmed=True,\
                       report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('report__service_in_date')
            
            if queryset:
                obj['schedules'] = queryset
                staff_obj_list.append(obj)

        #給与出力用ファイル生成
        achieve = commissionemployee_achieve_list(staff_obj_list,year,month)

        #wb = openpyxl.load_workbook('aggregates/monthly_employee.xlsx')
        wb = openpyxl.Workbook()
        sheet = wb.active

        #罫線
        side   = openpyxl.styles.borders.Side(style='thin', color='000000')
        border = openpyxl.styles.borders.Border(top=side, bottom=side, left=side, right=side)
        #背景色
        fill   = openpyxl.styles.PatternFill(patternType='solid', fgColor='d3d3d3')
        fill_for_input = openpyxl.styles.PatternFill(patternType='solid', fgColor='FFFF00')

        font = openpyxl.styles.Font(name='BIZ UDゴシック')

        for staff_data in achieve:

            sheet_name = "R" + str(year-2018) + "." + str(month) + " " + staff_data['staff_name']

            #シートの存在を確認
            is_sheet = False
            for ws in wb.worksheets:
                if ws.title == sheet_name:
                    is_sheet = True
                    break
            
            

            #シートが存在していなければ作成
            if not is_sheet: 
                wb.create_sheet(title=sheet_name,index=-1)
                ws = wb[sheet_name]
                ws.sheet_view.showGridLines = False #目盛り線を消す
                

                print_start = "A1"
                row_height = 24

                #列の幅を調整
                ws.column_dimensions['A'].width = 2.5
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['F'].width = 22
                ws.column_dimensions['G'].width = 14
                ws.column_dimensions['H'].width = 7
                ws.column_dimensions['I'].width = 10
                ws.column_dimensions['J'].width = 10
                ws.column_dimensions['k'].width = 8

                ws.print_title_rows = '1:2' #印刷時ヘッダー名前と行タイトル固定

            
                row = ws.max_row
                ws.row_dimensions[row].height = 30  #行の高さ
                ws.cell(row,2,value=str(year) + "年" + str(month) + "月  " + staff_data['staff_name'] + " 様")
                ws.cell(row,2).font = openpyxl.styles.fonts.Font(size=16)
                ws.cell(row,2).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')

                row +=1
                ws.row_dimensions[row].height = row_height #行の高さ
                ws.cell(row,2,value="日付")
                ws.cell(row,3,value="合計時間")
                ws.cell(row,4,value="利用者")
                ws.cell(row,5,value="実施分数")
                ws.cell(row,6,value="サービス")
                ws.cell(row,7,value="実施時間")
                ws.cell(row,8,value="研修")
                ws.cell(row,9,value="単価")
                ws.cell(row,10,value="日額")
                if staff_data['pay_bike']:
                    ws.cell(row,11,value="バイク代")
                else:
                    ws.merge_cells(ws.cell(row=row,column=10).coordinate + ":" + ws.cell(row=row,column=11).coordinate)

                #センターリング・罫線・背景色
                for r in  ws.iter_rows(min_row=row, min_col=2, max_row=row, max_col=11):
                    for c in r:
                        c.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                        ws[c.coordinate].border = border
                        ws[c.coordinate].fill   = fill

                index = row+1
                start_row = index #合計値計算用
                for day,data in staff_data['days_data'].items():
                    if data['schedules']:
                        day_start_row = index
                        day_end_row   = index + len(data['schedules'])-1
                        ws.cell(index,2,value= str(day) + "日(" + data['week'] + ")")
                        ws.cell(index,2).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                        ws.cell(index,3, value= data['day_hour'])
                        ws.cell(index,3).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                        ws.cell(index,10,value= data['day_pay'])
                        ws.cell(index,10).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                        if staff_data['pay_bike']:
                            ws.cell(index,11,value= data['day_bike_cost'])
                            ws.cell(index,11).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')

                        #セル結合///////
                        #day
                        ws.merge_cells(ws.cell(row=index,column=2).coordinate + ":" + ws.cell(row=index+len(data['schedules'])-1,column=2).coordinate)
                        #hour
                        ws.merge_cells(ws.cell(row=index,column=3).coordinate + ":" + ws.cell(row=index+len(data['schedules'])-1,column=3).coordinate)
                        #日額
                        
                        if staff_data['pay_bike']:
                            ws.merge_cells(ws.cell(row=index,column=10).coordinate + ":" + ws.cell(row=index+len(data['schedules'])-1,column=10).coordinate)
                        else:
                            ws.merge_cells(ws.cell(row=index,column=10).coordinate + ":" + ws.cell(row=index+len(data['schedules'])-1,column=11).coordinate)
                        #bike
                        ws.merge_cells(ws.cell(row=index,column=11).coordinate + ":" + ws.cell(row=index+len(data['schedules'])-1,column=11).coordinate)

                        #sum式 & 書式/////////
                        #日額
                        ws.cell(index,10,value='=SUM(' + ws.cell(row=day_start_row,column=9).coordinate + ':' + ws.cell(row=day_end_row,column=9).coordinate + ')')
                        ws.cell(index,10).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                        #bike
                        ws.cell(index,10).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')

                        
                        for sche in data['schedules']:
                                                    
                            ws.row_dimensions[index].height = row_height #行の高さ

                            ws.cell(index,4,value=sche['careuser'].last_name + "　" + sche['careuser'].first_name + "　様")
                            ws.cell(index,4).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                            ws.cell(index,5,value=sche['real_minutes'])
                            ws.cell(index,5).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                            if sche['check_min']:ws.cell(index,5).font = openpyxl.styles.fonts.Font(color='FF0000')
                            ws.cell(index,6,value=sche['service'])
                            if sche['check_min']:ws.cell(index,6).font = openpyxl.styles.fonts.Font(color='FF0000')
                            ws.cell(index,6).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                            ws.cell(index,7,value=sche['s_in_time'] + "～" + sche['s_out_time'])
                            ws.cell(index,7).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                            if sche['kenshuu']: ws.cell(index,8,value="[研修]")
                            ws.cell(index,8).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                            ws.cell(index,9,value=sche['pay'])
                            ws.cell(index,9).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')

                            index += 1

                end_row = index-1 #合計値計算用
                #罫線
                for r in  ws.iter_rows(min_row=start_row, min_col=2, max_row=end_row, max_col=11):
                    for c in r:
                        ws[c.coordinate].border = border

                row = ws.max_row+1 

                ws.row_dimensions[row].height = row_height #行の高さ
                ws.cell(row,2,value='小計')
                ws.cell(row,2).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')                

                #hour計
                ws.cell(row,3,value='=SUM(' + ws.cell(row=start_row,column=3).coordinate + ':' + ws.cell(row=end_row,column=3).coordinate + ')')
                ws.cell(row,3).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')

                if staff_data['pay_bike']:
                    #日額計
                    ws.cell(row,10,value='=SUM(' + ws.cell(row=start_row,column=10).coordinate + ':' + ws.cell(row=end_row,column=10).coordinate + ')')
                    ws.cell(row,10).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                    #bike計
                    ws.cell(row,11,value='=SUM(' + ws.cell(row=start_row,column=11).coordinate + ':' + ws.cell(row=end_row,column=11).coordinate + ')')
                    ws.cell(row,11).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                else:
                    #日額計
                    ws.cell(row,10,value='=SUM(' + ws.cell(row=start_row,column=10).coordinate + ':' + ws.cell(row=end_row,column=10).coordinate + ')')
                    ws.merge_cells(ws.cell(row,10).coordinate + ":" + ws.cell(row,11).coordinate)
                    ws.cell(row,10).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                   
                ws.row_dimensions[row+2].height = row_height #行の高さ
                ws.cell(row+2,9,value='合計')
                ws.cell(row+2,9).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
                ws.cell(row+2,9).fill   = fill
                ws.cell(row+2,9).border = border

                ws.merge_cells(ws.cell(row=row+2,column=10).coordinate + ":" + ws.cell(row=row+2,column=11).coordinate)
                ws.cell(row+2,10,value='=SUM(' + ws.cell(row=row,column=10).coordinate + ':' + ws.cell(row=row,column=11).coordinate + ')')
                ws.cell(row+2,10).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
                ws.cell(row+2,10).border = border
                ws.cell(row+2,11).border = border

                row = ws.max_row+2 

                #印刷範囲
                print_end = ws.cell(row=row,column=11).coordinate
                ws.print_area = print_start + ":" + print_end            
                ws.page_setup.fitToWidth  = True
                ws.page_setup.fitToHeight = False
                ws.sheet_properties.pageSetUpPr.fitToPage = True

        #font
        #for row in ws:
        #    for cell in row:
        #        ws[cell.coordinate].font = font

        #出力
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=monthly_commission.xlsx'
        # データの書き込みを行なったExcelファイルを保存する
        #wb.save('aggregates/monthly_employee.xlsx')
        wb.save(response)

        # 生成したHttpResponseをreturnする
        return response 
    else:
        return Http404

def commissionemployee_achieve_list(staff_obj_list,year,month):

    archive = []

    for s in staff_obj_list:
        #一カ月の日数を取得(最終日のみ)
        days = calendar.monthrange(year,month)[1]
        obj_by_staff ={}
        obj_by_staff['staff_name'] = s['staff'].last_name + " " + s['staff'].first_name
        obj_by_staff['month_hour'] = 0
        obj_by_staff['pay_bike'] = s['staff'].pay_bike
        obj_by_staff['month_pay'] = 0
        obj_by_staff['month_bike_cost'] = 0

        staff_days_data = {}
        for day in range(days):
            # {1日のdatetime: 1日のスケジュール全て, 2日のdatetime: 2日の全て...}のような辞書を作る
            staff_days_data[day+1] = {}
            staff_days_data[day+1]['week'] = jpweek(make_aware(datetime.datetime(year,month,day+1,0,0)))
            staff_days_data[day+1]['schedules'] = []
            staff_days_data[day+1]['day_hour'] = 0
            staff_days_data[day+1]['day_bike_cost'] = 0
            staff_days_data[day+1]['day_pay'] = 0 


        for sche in s['schedules']:
            d ={}
            s_in_date  = localtime(sche.report.service_in_date)
            s_out_date = localtime(sche.report.service_out_date)
            
            d['real_minutes']  = math.floor((s_out_date - s_in_date).total_seconds()/60)
            d['real_hour']     = math.ceil(d['real_minutes']/15)*0.25
            d['s_in_time']  = str(s_in_date.hour).zfill(2)  + ":" + str(s_in_date.minute).zfill(2)
            d['s_out_time'] = str(s_out_date.hour).zfill(2) + ":" + str(s_out_date.minute).zfill(2)
            d['s_in_time_datetime']  = s_in_date
            d['s_out_time_datetime'] = s_out_date

            
            d['careuser'] = sche.careuser
            d['service'] = ""
            if sche.service.kind==0:d['service'] += "[介護]"
            elif sche.service.kind==1:d['service'] += "[障害]"
            d['service']  += sche.service.title

            d['service_minutes']  = sche.service.time
            d['service_hour']     = math.ceil(sche.service.time/15)*0.25

            d['check_min'] = True if d['real_minutes'] != d['service_minutes'] else False

            #同行チェック
            d['kenshuu'] = False
            if sche.tr_staff1 == s['staff'] or sche.tr_staff2 == s['staff'] or sche.tr_staff3 == s['staff'] or sche.tr_staff4 == s['staff']:
                d['kenshuu'] = True

            #備考に入力があれば付記。
            d['biko'] = ""
            if sche.biko:d['biko'] += sche.biko
            if sche.report.communicate:
                if d['biko']:d['biko'] += "　"
                d['biko'] += sche.report.communicate
            
            #時間計算///////////////////////////////////////////////////////////////////////
            #実質時間または規定時間を計算適用時間とする。
            """
            if d['real_hour'] > d['service_hour']:
                d['adopt_hour'] = d['real_hour']
            else:
                d['adopt_hour'] = d['real_hour']
            #合計時間に加算する。
            staff_days_data[s_in_date.day]['day_hour'] += d['adopt_hour']
            obj_by_staff['month_hour'] += d['adopt_hour']
            """
            staff_days_data[s_in_date.day]['day_hour'] += d['real_hour']
            obj_by_staff['month_hour'] += d['real_hour']


            #支給額計算
            pay_by_sche = get_pay(sche,d['kenshuu'],s)
            d['pay'] = pay_by_sche
            staff_days_data[s_in_date.day]['day_pay'] += pay_by_sche
            obj_by_staff['month_pay']         += pay_by_sche

            staff_days_data[s_in_date.day]['schedules'].append(d)

        obj_by_staff['days_data'] = staff_days_data

        if s['staff'].pay_bike:
            #バイク代を加算
            #同一careuserの続いているスケジュールは加算しない。
            for day in range(len(obj_by_staff['days_data'])):
                #まず日毎の全スケジュール数を取得
                bike_cnt = len(obj_by_staff['days_data'][day+1]['schedules'])

                #前日または当日に同一利用者で繋がるスケジュールがあれば減算する
                for sche in obj_by_staff['days_data'][day+1]['schedules']:
                    #前日をチェック
                    #1日で0時0分からのサービスの場合はDBでチェック                    
                    if day == 0 and sche['s_in_time_datetime'].hour==0 and sche['s_in_time_datetime'].minute==0:
                        queryset = Schedule.objects.select_related('report','careuser').filter(search_staff_tr_query(s['staff']),careuser=sche['careuser'],report__careuser_confirmed=True,\
                        report__service_out_date=sche['s_in_time_datetime'],cancel_flg=False)
                        if queryset:
                           bike_cnt -= 1
                    #1日以外の0時0分スタートは前日をチェック
                    elif day >0 and sche['s_in_time_datetime'].hour==0 and sche['s_in_time_datetime'].minute==0:
                        if obj_by_staff['days_data'][day]['schedules']:
                            for sc in obj_by_staff['days_data'][day]['schedules']:
                                if sc['s_out_time_datetime'] == sche['s_in_time_datetime'] and sc['careuser'] == sche['careuser']:
                                    bike_cnt -= 1
                    #上記以外はdictにてチェック
                    else:
                        if obj_by_staff['days_data'][day+1]['schedules']:
                            for sc in obj_by_staff['days_data'][day+1]['schedules']:
                                if sc['s_out_time_datetime'] == sche['s_in_time_datetime'] and sc['careuser'] == sche['careuser']:
                                    bike_cnt -= 1

                if bike_cnt == 0:
                    obj_by_staff['days_data'][day+1]['day_bike_cost'] = 0
                elif bike_cnt == 1:
                    obj_by_staff['days_data'][day+1]['day_bike_cost'] = 100
                elif bike_cnt > 1:
                    if s['staff'].last_name == "岡田" and s['staff'].first_name == "真紀":#岡田真紀さんのみバイク代の上限なし
                        obj_by_staff['days_data'][day+1]['day_bike_cost'] = 100 * bike_cnt
                    else:
                        obj_by_staff['days_data'][day+1]['day_bike_cost'] = 200

                obj_by_staff['month_bike_cost'] += obj_by_staff['days_data'][day+1]['day_bike_cost']


        #支給合計額
        obj_by_staff['month_total_pay'] = obj_by_staff['month_pay'] + obj_by_staff['month_bike_cost']

        archive.append(obj_by_staff)

    return archive

def get_pay(sche,kenshuu,target_staff):

    #まず昼間のギャラを取得
    if not kenshuu:
        if sche.service.kind==0:#介護保険
            if "身体" in sche.service.bill_title and "生活" in sche.service.bill_title:
                if "身体1" in sche.service.bill_title:   sin=30                
                elif "身体2" in sche.service.bill_title: sin=60
                elif "身体3" in sche.service.bill_title: sin=90
                elif "身体4" in sche.service.bill_title: sin=120

                if "生活1" in sche.service.bill_title:   sei=30                
                elif "生活2" in sche.service.bill_title: sei=60
                elif "生活3" in sche.service.bill_title: sei=90
                elif "生活4" in sche.service.bill_title: sei=120
                pay = gur_sinsei(sin,sei,sche)

            elif "身体" in sche.service.bill_title:
                if "身体介護01" in sche.service.bill_title:  min=30  
                elif "身体介護1" in sche.service.bill_title: min=30                
                elif "身体介護2" in sche.service.bill_title: min=60
                elif "身体介護3" in sche.service.bill_title: min=90
                elif "身体介護4" in sche.service.bill_title: min=120
                pay = gur_sintai(min,sche,target_staff)

            elif "生活" in sche.service.bill_title:
                if "生活援助1" in sche.service.bill_title:   min=30                
                elif "生活援助2" in sche.service.bill_title: min=45
                elif "生活援助3" in sche.service.bill_title: min=60

                pay = gur_seikatu(min,sche)

        elif sche.service.kind==1:#障害
            if "身体" in sche.service.bill_title and "家事" in sche.service.bill_title:
                t = sche.service.bill_title.split('/')
                sin = int(re.sub(r"\D", "", t[0]))
                sei = int(re.sub(r"\D", "", t[1]))
                min = sin+sei
                pay = gur_sinsei(sin,sei,sche)

            elif "身体" in sche.service.bill_title:
                min = int(re.sub(r"\D", "", sche.service.bill_title))
                pay = gur_sintai(min,sche,target_staff)

            elif "家事" in sche.service.bill_title:
                min = int(re.sub(r"\D", "", sche.service.bill_title))
                pay = gur_seikatu(min,sche)

            elif "重度" in sche.service.bill_title:
                min = int(re.sub(r"\D", "", sche.service.bill_title))
                pay = gur_juudo(min,sche)

            elif "通院" in sche.service.bill_title:
                min = int(re.sub(r"\D", "", sche.service.bill_title))
                pay = gur_tuuin(min,sche) if "身有" in sche.service.bill_title else gur_seikatu(min,sche) #身体なしは生活の料金と同じ

        elif sche.service.kind==2:#移動支援
            min = int(re.sub(r"\D", "", sche.service.bill_title))
            pay = gur_idou_ari(min,sche) if "身有" in sche.service.bill_title else gur_idou_nasi(min,sche)

        elif sche.service.kind==3:#総合事業
            min = int(re.sub(r"\D", "", sche.service.bill_title))
            pay = gur_yobou(min,sche)

        elif sche.service.kind==4:#同行援護
            min = int(re.sub(r"\D", "", sche.service.bill_title))
            pay = gur_doukou(min,sche)

        elif sche.service.kind==5:#自費　night,midnight加算は付けない
            min = int(re.sub(r"\D", "", sche.service.bill_title))
            pay = gur_jihi(min,sche)

    else:
        #研修はnight,midnight加算は付けない
        s_in_datetime  = localtime(sche.report.service_in_date)
        s_out_datetime = localtime(sche.report.service_out_date)

        if sche.service.mix_items:
            pay = gur_kenshuu(sche.report.in_time_main + sche.report.in_time_sub)
        else:
            min = math.floor((s_out_datetime - s_in_datetime).total_seconds()/60)
            pay = gur_kenshuu(min)

    return pay


def gur_sintai(min,sche,target_staff):

    def daytime_gur(minutes):
        if minutes<20:
            rt = 650
        elif minutes <=30:
            rt = 800
        elif minutes <=60:
            rt = 1600
        else:
            add30 = math.ceil(max(0,(minutes-60))/30)
            rt = 1600 + 600*add30
        return rt

    s_in_datetime  = localtime(sche.report.service_in_date)
    s_out_datetime = s_in_datetime + datetime.timedelta(minutes=min)
    oc0  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=0,minute=0,second=0)))
    oc6  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=6,minute=0,second=0)))
    oc8  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=8,minute=0,second=0)))
    oc18 = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=18,minute=0,second=0)))
    oc22 = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=22,minute=0,second=0)))
    oc24 = oc0 + relativedelta(days=1)

    if sche.service.kind==0:#介護保険は開始時刻に合わせて一律
        if (s_in_datetime >= oc0 and  s_in_datetime < oc6) or (s_in_datetime >= oc22 and  s_in_datetime < oc24):
            gur = math.floor(daytime_gur(min)*1.75)
        elif (s_in_datetime >= oc6 and  s_in_datetime < oc8) or (s_in_datetime >= oc18 and  s_in_datetime < oc22):
            gur = math.floor(daytime_gur(min)*1.25)
        else:
            gur = math.floor(daytime_gur(min))
    elif sche.service.kind==1:#障害は時間によって変動
        times = get_separate_times(min,s_in_datetime)
        day_pay = daytime_gur(math.floor((times['s_out_date'] - times['s_in_date']).total_seconds()/60/30)*30)
        add_pay = 0#夜間・深夜料金
        
        if times['time_0to6_start']  :add_pay += 0.75 * daytime_gur(math.floor((times['time_0to6_end']   - times['time_0to6_start']  ).total_seconds()/60/30)*30)
        if times['time_6to8_start']  :add_pay += 0.25 * daytime_gur(math.floor((times['time_6to8_end']   - times['time_6to8_start']  ).total_seconds()/60/30)*30)
        #岡田真紀さんの木戸さんの夜間分のみ深夜料金にて計算
        if times['time_18to22_start'] and times['time_22to24_start'] and target_staff['staff'].last_name == "岡田" and target_staff['staff'].first_name == "真紀" and sche.careuser.last_name == "木戸" and sche.careuser.first_name == "功":
            add_pay += 0.75 * daytime_gur(math.floor((times['time_22to24_end'] - times['time_18to22_start']).total_seconds()/60/30)*30)
        else:
            if times['time_18to22_start']:add_pay += 0.25 * daytime_gur(math.floor((times['time_18to22_end'] - times['time_18to22_start']).total_seconds()/60/30)*30)
            if times['time_22to24_start']:add_pay += 0.75 * daytime_gur(math.floor((times['time_22to24_end'] - times['time_22to24_start']).total_seconds()/60/30)*30)
        gur = math.floor(day_pay +add_pay)
            
    return gur

def gur_seikatu(min,sche):
    def daytime_gur(minutes):
        if minutes <=30:
            rt = 650
        elif minutes <=45:
            if sche.service.bill_title == "生活援助2":
                rt = 1100
            else:
                rt = 1000
        elif minutes <=60:
            rt = 1300
        else:
            add30 = math.ceil(max(0,(minutes-60))/30)
            rt = 1300 + 600*add30
        return rt

    s_in_datetime  = localtime(sche.report.service_in_date)
    s_out_datetime = s_in_datetime + datetime.timedelta(minutes=min)
    oc0  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=0,minute=0,second=0)))
    oc6  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=6,minute=0,second=0)))
    oc8  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=8,minute=0,second=0)))
    oc18 = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=18,minute=0,second=0)))
    oc22 = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=22,minute=0,second=0)))
    oc24 = oc0 + relativedelta(days=1)

    if sche.service.kind==0:#介護保険は開始時刻に合わせて一律
        if (s_in_datetime >= oc0 and s_in_datetime < oc6) or (s_in_datetime >= oc22 and s_in_datetime < oc24):
            gur = math.floor(daytime_gur(min)*1.75)
        elif (s_in_datetime >= oc6 and s_in_datetime < oc8) or (s_in_datetime >= oc18 and s_in_datetime < oc22):
            gur = math.floor(daytime_gur(min)*1.25)
        else:
            gur = math.floor(daytime_gur(min))
    elif sche.service.kind==1:#障害は時間によって変動
        times = get_separate_times(min,s_in_datetime)
        day_pay = daytime_gur(math.floor((times['s_out_date'] - times['s_in_date']).total_seconds()/60/30)*30)
        add_pay = 0#夜間・深夜料金
        if times['time_0to6_start']  :add_pay += 0.75 * daytime_gur(math.floor((times['time_0to6_end']   - times['time_0to6_start']  ).total_seconds()/60/30)*30)
        if times['time_6to8_start']  :add_pay += 0.25 * daytime_gur(math.floor((times['time_6to8_end']   - times['time_6to8_start']  ).total_seconds()/60/30)*30)
        if times['time_18to22_start']:add_pay += 0.25 * daytime_gur(math.floor((times['time_18to22_end'] - times['time_18to22_start']).total_seconds()/60/30)*30)
        if times['time_22to24_start']:add_pay += 0.75 * daytime_gur(math.floor((times['time_22to24_end'] - times['time_22to24_start']).total_seconds()/60/30)*30)
        gur = math.floor(day_pay +add_pay)
            
    return gur

def gur_sinsei(sin,sei,sche):    
    def sin_daytime_gur(sin_min):
        gur_sin = 800 if sin_min <=30 else 1600 if sin_min <=60 else 1600 + 600*math.ceil((sin_min-60)/30)
        return gur_sin

    def sei_daytime_gur(sei_min):
        gur_sei = 600 if sei_min <=30 else 600 * math.ceil(sei_min/30)
        return gur_sei

    s_in_datetime  = localtime(sche.report.service_in_date)
    oc0  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=0,minute=0,second=0)))
    oc6  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=6,minute=0,second=0)))
    oc8  = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=8,minute=0,second=0)))
    oc18 = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=18,minute=0,second=0)))
    oc22 = make_aware(datetime.datetime.combine(s_in_datetime.date(),datetime.time(hour=22,minute=0,second=0)))
    oc24 = oc0 + relativedelta(days=1)

    if sche.service.kind==0:#介護保険は開始時刻に合わせて一律
        if (s_in_datetime >= oc0 and  s_in_datetime < oc6) or (s_in_datetime >= oc22 and  s_in_datetime < oc24):
            gur = math.floor(sin_daytime_gur(sin)*1.75) + math.floor(sei_daytime_gur(sei)*1.75)
        elif (s_in_datetime >= oc6 and  s_in_datetime < oc8) or (s_in_datetime >= oc18 and  s_in_datetime < oc22):
            gur = math.floor(sin_daytime_gur(sin)*1.25) + math.floor(sei_daytime_gur(sei)*1.25)
        else:
            gur = math.floor(sin_daytime_gur(sin)) + math.floor(sei_daytime_gur(sei))
    elif sche.service.kind==1:#障害は時間によって変動
        if not sche.report.mix_reverse:
            sin_s_in_datetime = s_in_datetime
            sei_s_in_datetime = sin_s_in_datetime + datetime.timedelta(minutes=sche.service.in_time_main)
        else:
            sei_s_in_datetime = s_in_datetime
            sin_s_in_datetime = sei_s_in_datetime + datetime.timedelta(minutes=sche.service.in_time_sub)

        sin_times = get_separate_times(sin,sin_s_in_datetime)
        sin_day_pay = sin_daytime_gur(math.floor((sin_times['s_out_date'] - sin_times['s_in_date']).total_seconds()/60/30)*30)
        sin_add_pay = 0#夜間・深夜料金
        if sin_times['time_0to6_start']  :sin_add_pay += 0.75 * sin_daytime_gur(math.floor((sin_times['time_0to6_end']   - sin_times['time_0to6_start']  ).total_seconds()/60/30)*30)
        if sin_times['time_6to8_start']  :sin_add_pay += 0.25 * sin_daytime_gur(math.floor((sin_times['time_6to8_end']   - sin_times['time_6to8_start']  ).total_seconds()/60/30)*30)
        if sin_times['time_18to22_start']:sin_add_pay += 0.25 * sin_daytime_gur(math.floor((sin_times['time_18to22_end'] - sin_times['time_18to22_start']).total_seconds()/60/30)*30)
        if sin_times['time_22to24_start']:sin_add_pay += 0.75 * sin_daytime_gur(math.floor((sin_times['time_22to24_end'] - sin_times['time_22to24_start']).total_seconds()/60/30)*30)
        
        sei_times = get_separate_times(sei,sei_s_in_datetime)
        sei_day_pay = sei_daytime_gur(math.floor((sei_times['s_out_date'] - sei_times['s_in_date']).total_seconds()/60/30)*30)
        sei_add_pay = 0#夜間・深夜料金
        if sei_times['time_0to6_start']  :sei_add_pay += 0.75 * sei_daytime_gur(math.floor((sei_times['time_0to6_end']   - sei_times['time_0to6_start']  ).total_seconds()/60/30)*30)
        if sei_times['time_6to8_start']  :sei_add_pay += 0.25 * sei_daytime_gur(math.floor((sei_times['time_6to8_end']   - sei_times['time_6to8_start']  ).total_seconds()/60/30)*30)
        if sei_times['time_18to22_start']:sei_add_pay += 0.25 * sei_daytime_gur(math.floor((sei_times['time_18to22_end'] - sei_times['time_18to22_start']).total_seconds()/60/30)*30)
        if sei_times['time_22to24_start']:sei_add_pay += 0.75 * sei_daytime_gur(math.floor((sei_times['time_22to24_end'] - sei_times['time_22to24_start']).total_seconds()/60/30)*30)

        gur = math.floor(sin_day_pay + sin_add_pay + sei_day_pay + sei_add_pay)

    return gur

def gur_tuuin(min,sche):
    def daytime_gur(minutes):
        if minutes <=30:
            rt = 800
        elif minutes <=60:
            rt = 1600
        else:
            add30 = math.ceil(max(0,(minutes-60))/30)
            rt = 1600 + 600*add30
        return rt

    s_in_datetime  = localtime(sche.report.service_in_date)
    times = get_separate_times(min,s_in_datetime)
    day_pay = daytime_gur(math.floor((times['s_out_date'] - times['s_in_date']).total_seconds()/60/30)*30)
    add_pay = 0#夜間・深夜料金
    
    if times['time_0to6_start']  :add_pay += 0.75 * daytime_gur(math.floor((times['time_0to6_end']   - times['time_0to6_start']  ).total_seconds()/60/30)*30)
    if times['time_6to8_start']  :add_pay += 0.25 * daytime_gur(math.floor((times['time_6to8_end']   - times['time_6to8_start']  ).total_seconds()/60/30)*30)
    if times['time_18to22_start']:add_pay += 0.25 * daytime_gur(math.floor((times['time_18to22_end'] - times['time_18to22_start']).total_seconds()/60/30)*30)
    if times['time_22to24_start']:add_pay += 0.75 * daytime_gur(math.floor((times['time_22to24_end'] - times['time_22to24_start']).total_seconds()/60/30)*30)
    gur = math.floor(day_pay +add_pay)
            
    return gur

def gur_doukou(min,sche):
    price_30min = 600
    s_in_datetime  = localtime(sche.report.service_in_date)
    return get_gur(price_30min,get_separate_times(min,s_in_datetime))

def gur_juudo(min,sche):
    price_30min = 600
    s_in_datetime  = localtime(sche.report.service_in_date)

    times = get_separate_times(min,s_in_datetime)

    if sche.careuser.last_name == "阪本" and sche.careuser.first_name == "毅"\
            and ((times['time_0to6_start'] and times['time_0to6_start'].hour==4 and times['time_0to6_start'].minute==0 and times['time_0to6_end'].hour==6 and times['time_0to6_end'].minute==0)\
            or (times['time_22to24_start'] and times['time_22to24_start'].hour==22 and times['time_22to24_start'].minute==0 and times['time_22to24_end'].hour==0 and times['time_22to24_end'].minute==0)):
           pay =4500
    else:
        pay = get_gur(price_30min,times)
    return pay

def gur_jihi(min,sche):
    if sche.careuser.last_name == "浅田" and sche.careuser.first_name == "真奈美":
        rt = math.floor(1600*min/60)
    else:
        rt = math.floor(20*min)
    return rt

def gur_yobou(min,sche):
    price_30min = 550
    s_in_datetime  = localtime(sche.report.service_in_date)
    return get_gur(price_30min,get_separate_times(min,s_in_datetime))

def gur_kenshuu(min):
    return math.floor(15.5*min)

def gur_idou_ari(min,sche):
    def daytime_gur(minutes):
        if minutes <=30:
            rt = 800
        elif minutes <=60:
            rt = 1500
        else:
            add30 = math.ceil(max(0,(minutes-60))/30)
            rt = 1500 + 550*add30
        return rt
    s_in_datetime  = localtime(sche.report.service_in_date)
    times = get_separate_times(min,s_in_datetime)
    day_pay = daytime_gur(math.floor((times['s_out_date'] - times['s_in_date']).total_seconds()/60/30)*30)
    add_pay = 0#夜間・深夜料金
    if times['time_0to6_start']  :add_pay += 0.75 * daytime_gur(math.floor((times['time_0to6_end']   - times['time_0to6_start']  ).total_seconds()/60/30)*30)
    if times['time_6to8_start']  :add_pay += 0.25 * daytime_gur(math.floor((times['time_6to8_end']   - times['time_6to8_start']  ).total_seconds()/60/30)*30)
    if times['time_18to22_start']:add_pay += 0.25 * daytime_gur(math.floor((times['time_18to22_end'] - times['time_18to22_start']).total_seconds()/60/30)*30)
    if times['time_22to24_start']:add_pay += 0.75 * daytime_gur(math.floor((times['time_22to24_end'] - times['time_22to24_start']).total_seconds()/60/30)*30)
    gur = day_pay +add_pay

    return gur

def gur_idou_nasi(min,sche):
    price_30min = 550
    s_in_datetime  = localtime(sche.report.service_in_date)
    return get_gur(price_30min,get_separate_times(min,s_in_datetime))

def get_separate_times(min,s_in_date):

    s_out_date = s_in_date + datetime.timedelta(minutes=min) #実施時間でなく計画サービスによる終了時刻

    #30分刻みとなるため、夜間、深夜の切り替えタイミングの時刻を取得　例 17:45開始の場合は18:15~夜間 17:46開始の場合は17:46~夜間 30分の過半数が夜間であれば夜間加算となる
    change_min = s_in_date.minute if s_in_date.minute>30 else s_in_date.minute + 30 if s_in_date.minute>0 and s_in_date.minute < 30 else 0
    
    if change_min <45:
        set_min = change_min - 30 if change_min>=30 else change_min
        
        change_time = {
            'midnight_0' :make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=0,minute=0,second=0))),
            'night_6'    :make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=6,minute=set_min,second=0))),
            'day_8'      :make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=8,minute=set_min,second=0))),
            'night_18'   :make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=18,minute=set_min,second=0))),
            'midnight_22':make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=22,minute=set_min,second=0))),
        }

    else:
        set_min = change_min
        
        change_time = {
            'midnight_0' :make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=0,minute=0,second=0))),
            'night_6'    :make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=5,minute=set_min,second=0))),
            'day_8'      :make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=7,minute=set_min,second=0))),
            'night_18'   :make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=17,minute=set_min,second=0))),
            'midnight_22':make_aware(datetime.datetime.combine(s_in_date.date(),datetime.time(hour=21,minute=set_min,second=0))),
        }
    #Noneで初期化
    time_0to6_start=time_0to6_end=time_6to8_start=time_6to8_end=time_8to18_start=time_8to18_end=time_18to22_start=time_18to22_end=time_22to24_start=time_22to24_end = None

    if s_in_date < change_time['night_6']:#スタートが6時以前の場合    
        time_0to6_start = s_in_date
        if s_out_date > change_time['midnight_22']:
            time_0to6_end   = time_6to8_start   = change_time['night_6']
            time_6to8_end   = time_8to18_start  = change_time['day_8']
            time_8to18_end  = time_18to22_start = change_time['night_18']
            time_18to22_end = time_22to24_start = change_time['midnight_22']
            time_22to24_end = s_out_date
        elif s_out_date > change_time['night_18']:
            time_0to6_end   = time_6to8_start   = change_time['night_6']
            time_6to8_end   = time_8to18_start  = change_time['day_8']
            time_8to18_end  = time_18to22_start = change_time['night_18']
            time_18to22_end   = s_out_date
        elif s_out_date > change_time['day_8']:
            time_0to6_end   = time_6to8_start   = change_time['night_6']
            time_6to8_end   = time_8to18_start  = change_time['day_8']
            time_8to18_end    = s_out_date
        elif s_out_date > change_time['night_6']:
            time_0to6_end   = time_6to8_start   = change_time['night_6']
            time_6to8_end     = s_out_date
        else:
            time_0to6_end     = s_out_date

    elif s_in_date < change_time['day_8']:#スタートが8時以前の場合
        time_6to8_start = s_in_date
        if s_out_date > change_time['midnight_22']:
            time_6to8_end   = time_8to18_start  =  change_time['day_8']
            time_8to18_end  = time_18to22_start = change_time['night_18']
            time_18to22_end = time_22to24_start = change_time['midnight_22']
            time_22to24_end   = s_out_date
        elif s_out_date > change_time['night_18']:
            time_6to8_end   = time_8to18_start  =  change_time['day_8']
            time_8to18_end  = time_18to22_start = change_time['night_18']
            time_18to22_end   = s_out_date
        elif s_out_date > change_time['day_8']:
            time_6to8_end   = time_8to18_start  =  change_time['day_8']
            time_8to18_end    = s_out_date
        else:
            time_6to8_end     = s_out_date

    elif s_in_date < change_time['night_18']:#スタートが18時以前の場合
        time_8to18_start = s_in_date
        if s_out_date > change_time['midnight_22']:
            time_8to18_end    = time_18to22_start = change_time['night_18']
            time_18to22_end   = time_22to24_start = change_time['midnight_22']
            time_22to24_end   = s_out_date
        elif s_out_date > change_time['night_18']:
            time_8to18_end    = time_18to22_start = change_time['night_18']
            time_18to22_end   = s_out_date
        else:
            time_8to18_end    = s_out_date

    elif s_in_date < change_time['midnight_22']:#スタートが22時以前の場合
        time_18to22_start = s_in_date
        if s_out_date > change_time['midnight_22']:
            time_18to22_end   = time_22to24_start = change_time['midnight_22']
            time_22to24_end   = s_out_date
        else:
            time_18to22_end   = s_out_date
    
    else:
        time_22to24_start = s_in_date
        time_22to24_end   = s_out_date

    ret_obj = {
        's_in_date'        :s_in_date,
        's_out_date'       :s_out_date,
        'time_0to6_start'  :time_0to6_start,
        'time_0to6_end'    : time_0to6_end,
        'time_6to8_start'  :time_6to8_start,
        'time_6to8_end'    : time_6to8_end,
        'time_8to18_start' :time_8to18_start,
        'time_8to18_end'   : time_8to18_end,
        'time_18to22_start':time_18to22_start,
        'time_18to22_end'  : time_18to22_end,
        'time_22to24_start': time_22to24_start,
        'time_22to24_end'  : time_22to24_end
    }
    
    return ret_obj

def get_gur(day_rate,get_separate_times):#加算額を含めたトータル支給額を計算
    sp_time = get_separate_times
    gur0_6 = gur6_8 = gur8_18 = gur18_22 = gur22_24 = 0
    if sp_time['time_0to6_start']:gur0_6     = math.floor(day_rate * 1.75 * math.floor((sp_time['time_0to6_end'] - sp_time['time_0to6_start']).total_seconds()/60/30))
    if sp_time['time_6to8_start']:gur6_8     = math.floor(day_rate * 1.25 * math.floor((sp_time['time_6to8_end'] - sp_time['time_6to8_start']).total_seconds()/60/30))
    if sp_time['time_8to18_start']:gur8_18   = math.floor(day_rate * math.floor((sp_time['time_8to18_end'] - sp_time['time_8to18_start']).total_seconds()/60/30))
    if sp_time['time_18to22_start']:gur18_22 = math.floor(day_rate * 1.25 * math.floor((sp_time['time_18to22_end'] - sp_time['time_18to22_start']).total_seconds()/60/30))
    if sp_time['time_22to24_start']:gur22_24 = math.floor(day_rate * 1.75 * math.floor((sp_time['time_22to24_end'] - sp_time['time_22to24_start']).total_seconds()/60/30))

    return gur0_6 + gur6_8 + gur8_18 + gur18_22 + gur22_24


class WorktimeView(SuperUserRequiredMixin,TemplateView):
    model = Schedule
    template_name = "aggregates/worktime_month_select.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        year = self.kwargs.get('year')
        month= self.kwargs.get('month')
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)
        next_month     = this_month + relativedelta(months=1)
        before_month   = this_month - relativedelta(months=1)

        context['this_month']    = this_month
        context['next_month']    = next_month
        context['before_month']  = before_month

        return context

def worktime_export(request,year,month):
    
    if request.user.is_superuser:
        this_month     = make_aware(datetime.datetime(year,month,1))
        this_month_end = this_month + relativedelta(months=1) - datetime.timedelta(seconds=1)

        staffs = User.objects.filter(salary=2).order_by('-is_staff','last_kana','first_kana')
        staff_obj_list = []
        for staff in staffs:
            obj = {}
            obj['staff'] = staff

            queryset = Schedule.objects.select_related('report','careuser','service').filter(search_staff_tr_query(staff),report__careuser_confirmed=True,\
                       report__service_in_date__range=[this_month,this_month_end],cancel_flg=False).order_by('report__service_in_date')
            
            if queryset:
                obj['schedules'] = queryset
                staff_obj_list.append(obj)

        #給与出力用ファイル生成
        achieve = commissionemployee_achieve_list(staff_obj_list,year,month)

        wb = openpyxl.load_workbook('aggregates/monthly_worktime.xlsx')        
        ws = wb['origin']
        ws.title = "R" + str(year-2018) + "." + str(month)
        ws['E2'] = year
        ws['H2'] = month


        #日付・曜日を記入
        w_list = ['月', '火', '水', '木', '金', '土', '日']
        dates = calendar.monthrange(year, month)
        holiday_kenji   = [5,6]
        holiday_yuri    = [5,6]
        holiday_okura   = [0,3]
        holiday_tanaka  = [0,5]
        holiday_uematsu = [1,5]
        holiday_hirose  = [2,6]

        col = 5
        week = dates[0]
        for i in range(dates[1]):
            ws.cell(column=col, row=5, value=i+1)
            ws.cell(column=col, row=6, value=w_list[week])
            if week not in holiday_kenji   :ws.cell(column=col, row=7, value=8)
            if week not in holiday_okura   :ws.cell(column=col, row=8, value=8)
            if week not in holiday_yuri    :ws.cell(column=col, row=9, value=8)
            if week not in holiday_uematsu :ws.cell(column=col, row=10, value=8)
            if week not in holiday_hirose  :ws.cell(column=col, row=11, value=8)
            if week not in holiday_tanaka  :ws.cell(column=col, row=12, value=8)
            col += 1
            week = week+1 if week<6 else 0

        row=13        
        for staff_data in achieve:
            ws.cell(column=2, row=row, value='訪問介護職員')
            ws.cell(column=4, row=row, value=staff_data['staff_name'])
            col = 4
            for day,data in staff_data['days_data'].items():
                if data['day_hour']:
                    #0.5刻みにする
                    day_hour = math.floor(data['day_hour']*2)/2
                    ws.cell(column=col+int(day), row=row, value=day_hour)

            row += 1

        #印刷範囲
        print_start = "A1"
        print_end   = "AL56"
        ws.print_area = print_start + ":" + print_end
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth  = True
        ws.page_setup.fitToHeight = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        
        #出力
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=monthly_commission.xlsx'
        # データの書き込みを行なったExcelファイルを保存する
        #wb.save('aggregates/monthly_employee.xlsx')
        wb.save(response)

        # 生成したHttpResponseをreturnする
        return response 
    else:
        return Http404

